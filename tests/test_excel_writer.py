"""Tests for the Excel writer: normalization, dataframe build, pivot, output."""

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from conftest import import_excel_writer

xw = import_excel_writer()


class TestNormalizeValue:
    def test_plain_numbers_pass_through(self):
        assert xw._normalize_value(100) == 100
        assert xw._normalize_value(1.5) == 1.5

    def test_currency_string(self):
        assert xw._normalize_value("$8,400.50") == 8400.50

    def test_comma_grouped_string(self):
        assert xw._normalize_value("1,204,556") == 1204556

    def test_percentage_over_one_scaled(self):
        assert xw._normalize_value("71.2%") == pytest.approx(0.712)

    def test_percentage_under_one_kept(self):
        assert xw._normalize_value("0.5%") == 0.5

    def test_empty_and_none_are_nan(self):
        assert np.isnan(xw._normalize_value(""))
        assert np.isnan(xw._normalize_value(None))

    def test_non_numeric_string_returned(self):
        assert xw._normalize_value("Roku") == "Roku"


class TestNormalizeText:
    def test_all_caps_titled(self):
        assert xw._normalize_text("FAMOUS TATE") == "Famous Tate"

    def test_all_lower_titled(self):
        assert xw._normalize_text("famous tate") == "Famous Tate"

    def test_mixed_case_kept(self):
        assert xw._normalize_text("Famous TATE Inc") == "Famous TATE Inc"

    def test_none_and_empty(self):
        assert xw._normalize_text(None) == ""
        assert xw._normalize_text("") == ""


class TestBuildDataframe:
    def _row(self, campaign="Camp A", metric="Impressions", value=100.0):
        return {"client": "C", "campaign": campaign, "campaign_type": "",
                "source": "Architect", "metric_level": "", "metric_name": metric,
                "metric_value": value, "start_date": "2026-05-01",
                "end_date": "2026-05-31"}

    def test_empty_input(self):
        df = xw._build_dataframe([])
        assert df.empty
        assert list(df.columns) == xw.SCHEMA_COLUMNS

    def test_duplicate_keys_summed(self):
        df = xw._build_dataframe([self._row(value=100.0), self._row(value=50.0)])
        assert len(df) == 1
        assert df["metric_value"].iloc[0] == 150.0

    def test_distinct_keys_kept(self):
        df = xw._build_dataframe([self._row("Camp A"), self._row("Camp B")])
        assert len(df) == 2

    def test_text_values_deduplicated_keep_last(self):
        r1 = self._row(metric="Status", value="Active")
        r2 = self._row(metric="Status", value="Paused")
        df = xw._build_dataframe([r1, r2])
        assert len(df) == 1
        assert df["metric_value"].iloc[0] == "Paused"

    def test_existing_key_replaced_on_rerun(self):
        df = xw._build_dataframe([self._row(value=10.0)],
                                 existing_rows=[self._row(value=5.0)])
        assert df["metric_value"].iloc[0] == 10.0

    def test_ratio_duplicates_are_averaged(self):
        df = xw._build_dataframe([
            self._row(metric="CTR", value=0.10),
            self._row(metric="CTR", value=0.30),
        ])
        assert df["metric_value"].iloc[0] == pytest.approx(0.20)

    def test_metric_acronyms_are_preserved(self):
        parsed = [{
            "client_name": "Client",
            "campaign_name": "Campaign",
            "source_platform": "Platform",
            "campaign_metrics": {
                "Campaign|CTR": {
                    "campaign_name": "Campaign",
                    "universal_name": "CTR",
                    "value": 0.12,
                }
            },
        }]
        rows = xw._collect_rows(parsed)
        assert rows[0]["metric_name"] == "CTR"



class TestWriteToExcel:
    def test_end_to_end(self, tmp_path, client_data):
        out = str(tmp_path / "report.xlsx")
        result, vba_err = xw.write_to_excel(client_data, out, inject_vba=False)
        assert result == out and vba_err is None
        wb = load_workbook(out)
        assert set(wb.sheetnames) == \
            {"Search", "_SearchIndex", "_Config", "Unified Data"}
        assert wb["_SearchIndex"].sheet_state == "hidden"
        ws = wb["Unified Data"]
        headers = [c.value for c in ws[1]]
        assert headers == xw.SCHEMA_COLUMNS
        assert ws.max_row > 1
        # Search sheet is first and carries the defined names the VBA needs
        assert wb.sheetnames[0] == "Search"
        for nm in ("SearchCell", "DateStart", "DateEnd", "ResultsAnchor"):
            assert nm in wb.defined_names

    def test_search_index_and_config_content(self, tmp_path, client_data):
        out = str(tmp_path / "report.xlsx")
        xw.write_to_excel(client_data, out, inject_vba=False)
        wb = load_workbook(out)
        idx = [(r[0].value, r[1].value, r[2].value)
               for r in wb["_SearchIndex"].iter_rows(min_row=2)]
        kinds = {k for _, k, _ in idx}
        assert {"metric", "campaign"} <= kinds
        cfg = [(r[0].value, r[3].value)
               for r in wb["_Config"].iter_rows(min_row=2)]
        assert cfg, "config sheet empty"
        assert all(agg in ("sum", "avg") for _, agg in cfg)

    def test_empty_input(self, tmp_path):
        """Empty input must not crash; behavior documented here.

        Pre-hardening: silently writes an empty workbook (or fails to save).
        Post-hardening (Phase 4): raises ValueError with a clear message.
        """
        out = str(tmp_path / "empty.xlsx")
        try:
            xw.write_to_excel([], out, inject_vba=False)
        except (ValueError, IndexError):
            return  # acceptable: explicit rejection of empty input
        import os
        assert os.path.exists(out)

    def test_rerun_is_idempotent(self, tmp_path, client_data):
        """Writing twice must not duplicate rows or double metric values."""
        out = str(tmp_path / "rerun.xlsx")
        xw.write_to_excel(client_data, out, inject_vba=False)
        wb_first = load_workbook(out, data_only=True)
        first_ws = wb_first["Unified Data"]
        first = first_ws.max_row
        first_values = [tuple(cell.value for cell in row)
                        for row in first_ws.iter_rows(min_row=2)]

        xw.write_to_excel(client_data, out, inject_vba=False)
        wb_second = load_workbook(out, data_only=True)
        second_ws = wb_second["Unified Data"]
        second = second_ws.max_row
        second_values = [tuple(cell.value for cell in row)
                         for row in second_ws.iter_rows(min_row=2)]

        assert first == second
        assert first_values == second_values



class TestExcelSession:
    """The batch export reuses one Excel COM process (engine.excel_vba.
    ExcelSession). Where COM is unavailable the session must be a safe
    no-op and injection must keep its per-call fallback behavior."""

    @staticmethod
    def _com_available():
        try:
            import win32com.client  # noqa: F401
            return True
        except ImportError:
            return False

    def test_session_is_safe_noop_without_com(self):
        if self._com_available():
            pytest.skip("would launch a real Excel process")
        from engine.excel_vba import ExcelSession
        with ExcelSession() as session:
            assert session.app is None
        assert session.app is None  # exit is idempotent

    def test_inject_without_shared_app_keeps_fallback_reason(self, tmp_path):
        if self._com_available():
            pytest.skip("would launch a real Excel process")
        from engine.excel_vba import inject_search_vba
        xlsx = tmp_path / "r.xlsx"
        xlsx.write_bytes(b"x")

        final_path, reason = inject_search_vba(str(xlsx), app=None)

        assert final_path == str(xlsx)
        assert "pywin32" in reason
