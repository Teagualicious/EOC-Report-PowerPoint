"""Tests for the Search Dashboard (excel_search_dashboard) and the unified
workbook writer round-trip (v16.9)."""

import pandas as pd
import pytest
from openpyxl import load_workbook

from engine.excel_search_dashboard import (build_config_rows,
                                           build_search_index_rows,
                                           build_search_dashboard)


@pytest.fixture
def df():
    rows = [
        {"client": "C", "campaign": "Camp A", "campaign_type": "", "source": "P1",
         "metric_level": "zip:28167", "metric_name": "Impressions",
         "metric_value": 546, "start_date": "2026-05-01", "end_date": "2026-05-31"},
        {"client": "C", "campaign": "Camp A", "campaign_type": "", "source": "P1",
         "metric_level": "zip:28167", "metric_name": "Contributions",
         "metric_value": 8786, "start_date": "2026-05-01", "end_date": "2026-05-31"},
        {"client": "C", "campaign": "Camp B", "campaign_type": "", "source": "P1",
         "metric_level": "network:CNN", "metric_name": "Impressions",
         "metric_value": 100, "start_date": "2026-05-01", "end_date": "2026-05-31"},
        {"client": "C", "campaign": "Camp B", "campaign_type": "", "source": "P1",
         "metric_level": "date:2026-05-02", "metric_name": "Frequency",
         "metric_value": 2.5, "start_date": "2026-05-01", "end_date": "2026-05-31"},
    ]
    return pd.DataFrame(rows)


class TestSearchIndex:
    def test_level_values_searchable_by_value_part(self, df):
        """'28167' must resolve without typing the 'zip:' prefix."""
        rows = build_search_index_rows(df)
        hit = [r for r in rows if r[0] == "28167"]
        assert hit and hit[0][1] == "level_value"
        assert hit[0][2] == "zip:28167"          # canonical keeps the prefix
        assert "zip" in hit[0][3]                 # display shows the type

    def test_metric_aliases_included(self, df):
        rows = build_search_index_rows(df)
        canon = {r[0]: r[2] for r in rows if r[1] == "metric"}
        assert canon.get("impressions") == "Impressions"
        # dictionary alias resolves to the same canonical metric
        assert any(c == "Impressions" and t != "impressions"
                   for t, c in canon.items())

    def test_level_types_and_campaigns_present(self, df):
        rows = build_search_index_rows(df)
        kinds = {(r[1], r[2]) for r in rows}
        assert ("level_type", "zip") in kinds
        assert ("level_type", "network") in kinds
        assert ("campaign", "Camp A") in kinds

    def test_no_duplicate_terms(self, df):
        rows = build_search_index_rows(df)
        assert len(rows) == len({(r[0], r[1], r[2]) for r in rows})


class TestConfig:
    def test_ratio_metrics_marked_avg(self, df):
        cfg = {m: agg for m, _, _, agg in build_config_rows(df)}
        assert cfg["Frequency"] == "avg"
        assert cfg["Impressions"] == "sum"

    def test_preferred_order_first(self, df):
        order = [m for m, _, _, _ in build_config_rows(df)]
        assert order.index("Impressions") < order.index("Frequency")


class TestDashboardSheets:
    def test_sheets_layout_and_names(self, df, tmp_path):
        from openpyxl import Workbook
        wb = Workbook(); del wb["Sheet"]
        build_search_dashboard(wb, df)
        assert wb.sheetnames[0] == "Search"
        assert wb["_SearchIndex"].sheet_state == "hidden"
        assert wb["_Config"].sheet_state == "hidden"
        for name in ("SearchCell", "DateStart", "DateEnd", "ResultsAnchor"):
            assert name in wb.defined_names
        p = tmp_path / "t.xlsx"
        wb.save(p)  # must serialize cleanly
        assert load_workbook(p)["Search"]["B2"].value == "Campaign Data Search"

    def test_reexport_preserves_search_sheet(self, df, tmp_path):
        """xlsm re-export path: existing Search sheet must be left alone
        (its VBA code-behind would be stranded), hidden sheets rebuilt."""
        from openpyxl import Workbook
        wb = Workbook(); del wb["Sheet"]
        build_search_dashboard(wb, df)
        wb["Search"]["Z1"] = "sentinel"
        # simulate the writer's re-export: hidden sheets removed, Search kept
        del wb["_SearchIndex"]; del wb["_Config"]
        build_search_dashboard(wb, df)
        assert wb["Search"]["Z1"].value == "sentinel"
        assert "_SearchIndex" in wb.sheetnames


class TestWriterRoundTrip:
    def _client_data(self):
        return [{
            "client_name": "C", "campaign_type": "",
            "start_date": "2026-05-01", "end_date": "2026-05-31",
            "source_platform": "P1",
            "campaign_metrics": {
                "k": {"universal_name": "Impressions", "value": 546,
                       "campaign_name": "Camp A"}},
            "level_data": [
                {"metric_name": "Impressions", "metric_value": 546,
                 "metric_level": "zip:28167", "_campaign": "Camp A"}],
        }]

    def test_writes_search_plus_unified_only(self, tmp_path):
        from engine.excel_writer import write_to_excel
        out = str(tmp_path / "r.xlsx")
        path, vba_err = write_to_excel(self._client_data(), out)
        # Off-Windows the VBA injection must fail GRACEFULLY: xlsx intact
        assert path.endswith(".xlsx") and vba_err is not None
        wb = load_workbook(path)
        visible = [s for s in wb.sheetnames if wb[s].sheet_state != "hidden"]
        assert visible == ["Search", "Unified Data"]

    def test_reexport_appends_deduped(self, tmp_path):
        from engine.excel_writer import write_to_excel
        out = str(tmp_path / "r.xlsx")
        write_to_excel(self._client_data(), out)
        path, _ = write_to_excel(self._client_data(), out)  # same data again
        ws = load_workbook(path)["Unified Data"]
        assert ws.max_row == 3  # header + 2 unique rows, not doubled

    def test_reexport_over_xlsm_rebuilds_fresh_and_removes_stale(self, tmp_path):
        """Regression: the old re-export path edited the existing .xlsm in
        place with openpyxl, which cannot round-trip macro workbooks — the
        ActiveX search box was dropped and the VBA wiring corrupted, so the
        search broke on every re-export of the same period. The writer must
        instead harvest the rows, rebuild the workbook from scratch (full
        Search sheet + defined names), and never leave the stale macro
        workbook behind holding old data."""
        import os
        from engine.excel_writer import write_to_excel

        out = str(tmp_path / "r.xlsx")
        first_path, _ = write_to_excel(self._client_data(), out)
        # Simulate the Windows COM upgrade of the first export to .xlsm
        xlsm = str(tmp_path / "r.xlsm")
        os.rename(first_path, xlsm)

        path, _vba_err = write_to_excel(self._client_data(), out)

        assert path.endswith(".xlsx")     # injection unavailable off-Windows
        assert not os.path.exists(xlsm)   # stale macro workbook removed
        wb = load_workbook(path)
        assert wb.sheetnames[0] == "Search"   # fully rebuilt, not stripped
        for nm in ("SearchCell", "DateStart", "DateEnd", "ResultsAnchor"):
            assert nm in wb.defined_names
        assert {"_SearchIndex", "_Config"} <= set(wb.sheetnames)
        # Rows harvested from the .xlsm and merged — replaced, not doubled
        assert wb["Unified Data"].max_row == 3


class TestVbaContract:
    """Static checks tying modSearch.bas to the Python-built sheets. The
    VBA cannot run in CI, but its cell/grammar contract can be locked."""

    @staticmethod
    def _bas():
        import os
        from engine import excel_vba
        path = os.path.join(os.path.dirname(excel_vba.__file__),
                            "vba_src", "modSearch.bas")
        return open(path, encoding="utf-8").read()

    def test_row_constants_match_dashboard(self):
        from engine.excel_search_dashboard import (RESULTS_START_ROW,
                                                   SUGGEST_ROW)
        src = self._bas()
        assert f"Const RESULTS_ROW As Long = {RESULTS_START_ROW}" in src
        assert f"Const SUGGEST_ROW As Long = {SUGGEST_ROW}" in src

    def test_typed_pivot_contract(self):
        """The search renders ONLY the results table, with columns in the
        exact typed order (mColSeq) — no auto KPI/summary boxes. Locked
        after real reports showed unrequested aggregate boxes and
        dims-before-metrics reordering."""
        src = self._bas()
        assert "Auto KPI strip" not in src
        assert "mColSeq" in src
        assert "exact typed order" in src.lower() or \
               "EXACT typed order" in src


def test_copy_chip_present(df=None):
    import pandas as pd
    from openpyxl import Workbook
    from engine.excel_search_dashboard import (build_search_dashboard,
                                               COPY_CELL, RESULTS_START_ROW)
    frame = pd.DataFrame([{"campaign": "A", "metric_level": "zip:1",
                           "metric_name": "Impressions", "metric_value": 1,
                           "client": "C", "campaign_type": "", "source": "S",
                           "start_date": "", "end_date": ""}])
    wb = Workbook(); del wb["Sheet"]
    ws = build_search_dashboard(wb, frame)
    assert "Copy results" in str(ws[COPY_CELL].value)
    # the chip must live ABOVE the results area so table clears never eat it
    assert int(COPY_CELL[1:]) < RESULTS_START_ROW
