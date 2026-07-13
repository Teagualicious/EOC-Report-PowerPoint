"""End-to-end headless pipeline smoke test.

Simulates the full Quick Run flow without any UI:
parse files → apply platform config → filter by campaigns → write Excel →
read the workbook back and assert real values.
"""

from openpyxl import load_workbook

from conftest import import_data_pipeline, import_excel_writer

pipeline = import_data_pipeline()
xw = import_excel_writer()


def test_full_pipeline(tmp_path, sample_xlsx, sample_csv):
    from parsers.csv_parser import CSVParser
    from parsers.excel_parser import ExcelParser

    # 1. Parse raw exports
    architect = ExcelParser(sample_xlsx).parse()
    audiencetrak = CSVParser(sample_csv).parse()

    # 2. Apply a platform config to the Architect file (as Platform Setup would)
    config = {"sheets": [
        {"sheet_name": "Data by Devices", "columns": [
            {"column": "Order", "role": "campaign_id", "selected": True},
            {"column": "Device", "role": "level: device", "selected": True},
            {"column": "Impressions", "role": "metric", "selected": True},
            {"column": "100% Complete", "role": "metric", "selected": True},
        ]},
        {"sheet_name": "Data by Geo", "columns": [
            {"column": "Order", "role": "campaign_id", "selected": True},
            {"column": "Zip Code", "role": "level: zip", "selected": True},
            {"column": "Impressions", "role": "metric", "selected": True},
        ]},
    ]}
    pipeline.apply_platform_config(architect, config)
    architect["source_platform"] = "Architect"
    audiencetrak["source_platform"] = "AudienceTrak"

    all_parsed = [architect, audiencetrak]

    # 3. Campaign discovery (as the client wizard does it inline),
    #    then filter to one client's campaigns
    campaigns = set()
    for data in all_parsed:
        for mdata in data.get("campaign_metrics", {}).values():
            if mdata.get("campaign_name"): campaigns.add(mdata["campaign_name"])
        for ld in data.get("level_data", []):
            if ld.get("_campaign"): campaigns.add(ld["_campaign"])
    assert "Campaign A" in campaigns
    assert "Famous Tate - Streaming TV" in campaigns

    selected = ["Campaign A", "Campaign B"]
    client_data = pipeline.filter_data_by_campaigns(all_parsed, selected)
    for data in client_data:
        data["client_name"] = "Famous Tate"
        data["start_date"] = "2026-05-01"
        data["end_date"] = "2026-05-31"

    # 4. Export Excel
    out = str(tmp_path / "Famous_Tate_unified_report.xlsx")
    xw.write_to_excel(client_data, out, inject_vba=False)

    # 5. Assert on workbook contents
    wb = load_workbook(out)
    assert set(wb.sheetnames) == \
        {"Search", "_SearchIndex", "_Config", "Unified Data"}

    ws = wb["Unified Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}

    by_key = {}
    for r in rows:
        key = (r[idx["campaign"]], r[idx["metric_level"]], r[idx["metric_name"]])
        by_key[key] = r[idx["metric_value"]]

    # Roku impressions for Campaign A came through the platform config
    assert by_key[("Campaign A", "device:Roku", "Impressions")] == 20000
    # Zip data from the second sheet
    assert by_key[("Campaign A", "zip:33607", "Impressions")] == 22000
    # Client name normalized onto every row
    assert all(r[idx["client"]] == "Famous Tate" for r in rows)


def test_pptx_metrics_from_pipeline(sample_xlsx):
    """get_available_metrics consumes pipeline output for the template mapper."""
    from parsers.excel_parser import ExcelParser
    from conftest import import_metrics_catalog

    catalog = import_metrics_catalog()
    parsed = ExcelParser(sample_xlsx).parse()

    metrics, structured = catalog.get_available_metrics(
        [parsed], "Famous Tate", "2026-05-01", "2026-05-31")

    assert metrics["__client_name__"] == "Famous Tate"
    assert metrics["__start_month__"] == "May"
    assert metrics["__year__"] == "2026"
    # Impressions totals: device 80,000 vs zip 80,000 (equal) — total exists
    assert "Total Impressions" in metrics
    assert metrics["device:Roku:Impressions"] == 50000  # 20,000 + 30,000
    assert "device" in structured["breakdowns"]
