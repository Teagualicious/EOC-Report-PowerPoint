"""Search Dashboard sheet builder (openpyxl side).

Builds three sheets into the unified workbook:
  - "Search"        — the user-facing smart-search dashboard
  - "_SearchIndex"  — hidden: every searchable term -> (kind, canonical)
  - "_Config"       — hidden: metric column order / enabled / aggregation

The interactive behavior (parsing the comma-separated search, building the
results table, suggestion chips, calendar popups) lives in VBA injected by
engine/excel_vba.py at export time. This module only lays down data and
layout, so it is fully testable off-Windows.

Search grammar (mirrored by the VBA in excel_vba.py):
  - Terms are comma-separated; each resolves in priority order:
      level VALUE  (e.g. "28167", "CNN")   -> filter rows to that value
      level TYPE   (e.g. "zip", "network") -> rows become that type's values
      metric       (e.g. "impressions")    -> restrict/order metric columns
      campaign                              -> filter to that campaign
  - Metric term order = column order. No metric terms -> _Config order.
  - No level terms -> rows are campaigns (totals view).
"""

import logging

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from parsers.dictionary import load_dictionary, get_metric_aggregation

logger = logging.getLogger(__name__)

# Spectrum Reach palette
NAVY = "003057"          # brand navy
BLUE = "0271EB"          # brand bright blue (matches the PPT template accents)
LIGHT = "EEF2F7"
BANNER_TEXT = "FFFFFF"
HINT_ON_NAVY = "B9C9DB"
_title_font = Font(name="Calibri", size=16, bold=True, color=BANNER_TEXT)
_tag_font = Font(name="Calibri", size=9, bold=True, color=HINT_ON_NAVY)
_hint_font = Font(name="Calibri", size=9, italic=True, color="6B7A90")
_hint_navy_font = Font(name="Calibri", size=9, italic=True, color=HINT_ON_NAVY)
_label_font = Font(name="Calibri", size=10, bold=True, color=NAVY)
_search_font = Font(name="Calibri", size=11)
_btn_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_chip_font = Font(name="Calibri", size=11, bold=True, color=BLUE)
_banner_fill = PatternFill("solid", fgColor=NAVY)
_accent_fill = PatternFill("solid", fgColor=BLUE)
_btn_fill = PatternFill("solid", fgColor=BLUE)
_search_fill = PatternFill("solid", fgColor="FFFFFF")
_panel_fill = PatternFill("solid", fgColor=LIGHT)
_box_border = Border(*[Side(style="medium", color=NAVY)] * 4)


def _style_range(ws, ref, fill=None, border=None, font=None, align=None):
    """Apply styles to EVERY cell in a range — required for merged ranges,
    where styling only the top-left cell leaves the rest looking cut off.
    Accepts single-cell refs too."""
    area = ws[ref]
    if not isinstance(area, tuple):
        area = ((area,),)
    for row in area:
        for c in row:
            if fill is not None: c.fill = fill
            if border is not None: c.border = border
            if font is not None: c.font = font
            if align is not None: c.alignment = align

# Preferred metric ordering for the default column set
_PREFERRED = ["Impressions", "Completions", "100% Completions", "Contributions",
              "Clicks", "Reach", "Frequency", "Cost"]

# Cell addresses the VBA relies on (also published as defined names).
# CHANGING ANY OF THESE requires matching edits in vba_src/modSearch.bas
# (RESULTS_ROW / SUGGEST_ROW) and VBA_SHEET_SEARCH (button/chip addresses).
SEARCH_CELL = "B6"
BTN_CELL = "H6"
DATE_START_CELL = "J6"
DATE_END_CELL = "M6"
CHIP_START = "K6"
CHIP_END = "N6"
SUGGEST_ROW = 9
RESULTS_START_ROW = 12
COPY_CELL = "D11"


def _level_aliases(dictionary):
    """level prefix -> list of user-facing alias terms (incl. the prefix)."""
    out = {}
    for level_def in dictionary.get("level_definitions", {}).values():
        prefix = level_def.get("prefix", "")
        if not prefix:
            continue
        terms = {prefix}
        for alias in level_def.get("id_column_aliases", []):
            terms.add(alias)
        out[prefix] = sorted(terms)
    return out


def build_search_index_rows(df, dictionary=None):
    """Rows for _SearchIndex: (term_lower, kind, canonical, display).

    kind: metric | level_type | level_value | campaign
    canonical: what the VBA filters on —
      metric      -> exact metric_name as it appears in unified data
      level_type  -> the level prefix ("zip", "network", ...)
      level_value -> the full "prefix:value" metric_level string
      campaign    -> exact campaign name
    """
    dictionary = dictionary or load_dictionary()
    rows = []
    seen = set()

    def add(term, kind, canonical, display):
        key = (term.lower().strip(), kind, canonical)
        if term and key not in seen:
            seen.add(key)
            rows.append((term.lower().strip(), kind, canonical, display))

    # Metrics present in the data, plus dictionary aliases for them
    present_metrics = sorted(set(df["metric_name"].dropna()))
    alias_map = dictionary.get("metric_aliases", {})
    for metric in present_metrics:
        add(metric, "metric", metric, metric)
        for alias in alias_map.get(metric, []):
            add(alias, "metric", metric, metric)

    # Level types present in the data + their dictionary aliases
    level_alias = _level_aliases(dictionary)
    present_prefixes = sorted(set(
        str(lv).split(":", 1)[0] for lv in df["metric_level"].dropna()
        if ":" in str(lv)))
    for prefix in present_prefixes:
        for term in level_alias.get(prefix, [prefix]):
            add(term, "level_type", prefix, prefix.title())

    # Level values — searchable by the part AFTER the colon. Excel hands
    # date breakdowns over as datetimes; the user-facing term/display drop
    # the " 00:00:00" noise, but the CANONICAL string stays verbatim so the
    # VBA filter still matches the Unified Data cells exactly.
    for lv in sorted(set(str(v) for v in df["metric_level"].dropna()
                         if ":" in str(v))):
        prefix, value = lv.split(":", 1)
        if value:
            shown = value[:-9] if value.endswith(" 00:00:00") else value
            add(shown, "level_value", lv, f"{shown} ({prefix})")

    # Campaigns
    for camp in sorted(set(df["campaign"].dropna())):
        if camp:
            add(str(camp), "campaign", str(camp), str(camp))

    # Grouping keywords: let users say "campaign" / "client" to add row
    # dimensions ("campaign, zip" -> Campaign + Zip label columns)
    for term in ("campaign", "campaigns"):
        add(term, "row_group", "campaign", "Campaign (group rows)")
    for term in ("client", "clients"):
        add(term, "row_group", "client", "Client (group rows)")

    return rows


def build_config_rows(df):
    """Rows for _Config: (metric, enabled, order, agg)."""
    present = sorted(set(df["metric_name"].dropna()))
    ordered = [m for m in _PREFERRED if m in present] + \
              [m for m in present if m not in _PREFERRED]
    def agg_for(m):
        # Approved rule: rate = 100 * Σcompletions / Σimpressions
        if m in ("Completion Rate", "Completion Percent"):
            return "rate"
        return get_metric_aggregation(m)
    return [(m, True, i + 1, agg_for(m)) for i, m in enumerate(ordered)]


def build_search_dashboard(wb, df, dictionary=None):
    """Create Search + hidden _SearchIndex/_Config sheets in the workbook.

    If a Search sheet already exists, the layout is left alone and only the
    hidden data sheets and defined names are rebuilt. (Defensive only: the
    writer always passes a freshly built workbook — editing an existing
    macro workbook in place is exactly what used to break the search on
    re-export.)"""
    dictionary = dictionary or load_dictionary()

    if "Search" in wb.sheetnames:
        _build_hidden_sheets(wb, df, dictionary)
        _define_names(wb)
        return wb["Search"]

    # ── Search sheet (position 0) ──
    ws = wb.create_sheet("Search", 0)
    ws.sheet_properties.tabColor = BLUE
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    _style_range(ws, "A1:T80", fill=PatternFill("solid", fgColor="FFFFFF"))

    # Brand banner (rows 1-2): navy band, white title, tag line, blue rule
    _style_range(ws, "A1:N2", fill=_banner_fill)
    ws["B1"] = "SPECTRUM REACH  |  CLIENT DEVELOPMENT & PERFORMANCE"
    ws["B1"].font = _tag_font
    ws["B2"] = "Campaign Data Search"
    ws["B2"].font = _title_font
    # grammar hint lives under the search box
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 40
    _style_range(ws, "A3:N3", fill=_accent_fill)
    try:
        import os
        from openpyxl.drawing.image import Image as XlImage
        from config.paths import RESOURCE_DIR
        lp = os.path.join(RESOURCE_DIR, "assets", "logo_white.png")
        if os.path.exists(lp):
            img = XlImage(lp)
            img.height = 26
            img.width = int(1200 * 26 / 343)
            ws.add_image(img, "L1")
    except Exception:
        logger.debug("Banner logo not embedded", exc_info=True)
    ws.row_dimensions[3].height = 3

    # Labels row (small caps, above the inputs)
    _small_label = Font(name="Calibri", size=8, bold=True, color="6B7A90")
    ws.row_dimensions[4].height = 10          # air under the banner
    ws["B5"] = "SEARCH"
    ws["B5"].font = _small_label
    ws["J5"] = "START DATE"
    ws["J5"].font = _small_label
    ws["M5"] = "END DATE"
    ws["M5"].font = _small_label
    ws.row_dimensions[5].height = 12

    # Inputs row: wide search box, button, date boxes with calendar chips
    ws.merge_cells(f"{SEARCH_CELL}:F6")
    _style_range(ws, f"{SEARCH_CELL}:F6", fill=_search_fill,
                 border=_box_border, font=_search_font,
                 align=Alignment(vertical="center", indent=1, wrap_text=True))
    ws.row_dimensions[6].height = 44   # the INPUT row (was orphaned at default height since the row-6 move — the clipping root cause)

    ws[BTN_CELL] = "🔍  Search"
    _style_range(ws, BTN_CELL, fill=_btn_fill, font=_btn_font,
                 border=_box_border,
                 align=Alignment(horizontal="center", vertical="center"))

    for addr in (DATE_START_CELL, DATE_END_CELL):
        _style_range(ws, addr, fill=_search_fill, border=_box_border,
                     align=Alignment(vertical="center", horizontal="center"))
        ws[addr].number_format = "yyyy-mm-dd"
    for chip in (CHIP_START, CHIP_END):
        ws[chip] = "📅"
        _style_range(ws, chip, font=_chip_font,
                     align=Alignment(horizontal="center", vertical="center"))

    # Suggestions row (VBA writes clickable chips here)
    ws.row_dimensions[7].height = 8           # air under the inputs
    ws["B8"] = ("Type metrics, levels (zip, network...), values, or "
                "campaigns — commas separate; metric order = column order. "
                "Suggestions appear below — click one to add it:")
    ws["B8"].font = _hint_font
    ws.row_dimensions[8].height = 14
    ws.row_dimensions[SUGGEST_ROW].height = 18
    ws.row_dimensions[10].height = 10         # air before Results

    # Results header with a blue rule underneath + one-click Copy chip
    ws[f"B{RESULTS_START_ROW - 1}"] = "Results"
    ws[f"B{RESULTS_START_ROW - 1}"].font = Font(name="Calibri", size=12,
                                                 bold=True, color=NAVY)
    _style_range(ws, f"B{RESULTS_START_ROW - 1}:N{RESULTS_START_ROW - 1}",
                 border=Border(bottom=Side(style="thick", color=BLUE)))
    ws[COPY_CELL] = "📋 Copy results"
    _style_range(ws, COPY_CELL,
                 fill=PatternFill("solid", fgColor="DEE8F5"),
                 font=Font(name="Calibri", size=10, bold=True, color=NAVY),
                 align=Alignment(horizontal="center", vertical="center"),
                 border=Border(bottom=Side(style="thick", color=BLUE)))

    # No-macro notice — the VBA clears this on first run
    ws[f"B{RESULTS_START_ROW}"] = (
        "Enable macros (and Excel's 'Trust access to the VBA project object "
        "model' was needed at export) to use the live search. Data is in the "
        "Unified Data sheet.")
    ws[f"B{RESULTS_START_ROW}"].font = _hint_font

    # Airy grid: spacer columns G/I/L separate every element; ends ~col N
    ws.column_dimensions["A"].width = 2
    for col in "BCDEF":
        ws.column_dimensions[col].width = 17
    ws.column_dimensions["G"].width = 2       # spacer
    ws.column_dimensions["H"].width = 14      # Search button (no clipping)
    ws.column_dimensions["I"].width = 2       # spacer
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 4.5
    ws.column_dimensions["L"].width = 2       # spacer
    ws.column_dimensions["M"].width = 13
    ws.column_dimensions["N"].width = 4.5

    _build_hidden_sheets(wb, df, dictionary)
    _define_names(wb)
    return ws


def _build_hidden_sheets(wb, df, dictionary):
    # ── _SearchIndex (hidden) ──
    ws_i = wb.create_sheet("_SearchIndex")
    ws_i.sheet_state = "hidden"
    for ci, h in enumerate(["term", "kind", "canonical", "display"], 1):
        ws_i.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(build_search_index_rows(df, dictionary), 2):
        for ci, v in enumerate(row, 1):
            ws_i.cell(row=ri, column=ci, value=v)

    # ── _Config (hidden, user can unhide to customize defaults) ──
    ws_c = wb.create_sheet("_Config")
    ws_c.sheet_state = "hidden"
    for ci, h in enumerate(["metric", "enabled", "order", "agg"], 1):
        ws_c.cell(row=1, column=ci, value=h)
    for ri, (m, en, order, agg) in enumerate(build_config_rows(df), 2):
        ws_c.cell(row=ri, column=1, value=m)
        ws_c.cell(row=ri, column=2, value="Y" if en else "N")
        ws_c.cell(row=ri, column=3, value=order)
        ws_c.cell(row=ri, column=4, value=agg)



def _define_names(wb):
    """Defined names the VBA uses to locate everything."""
    for name, ref in [("SearchCell", f"Search!${SEARCH_CELL[0]}${SEARCH_CELL[1:]}"),
                      ("DateStart", f"Search!${DATE_START_CELL[0]}${DATE_START_CELL[1:]}"),
                      ("DateEnd", f"Search!${DATE_END_CELL[0]}${DATE_END_CELL[1:]}"),
                      ("ResultsAnchor", f"Search!$B${RESULTS_START_ROW}")]:
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
