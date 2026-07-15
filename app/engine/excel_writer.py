"""Excel report writer — pandas data pipeline + openpyxl styling.

Data flow:
1. Collect raw rows from parsed data into a pandas DataFrame (excel_utils)
2. Deduplicate by composite key, aggregating numeric values
3. Pivot for dashboard sheets
4. Write to Excel with openpyxl formatting
"""

import logging
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine.excel_utils import (
    SCHEMA_COLUMNS, _format_cell, _collect_rows, _build_dataframe,
    # re-exported for tests (pyflakes flags these; intentional):
    _normalize_value, _normalize_text,  # noqa: F401
)

logger = logging.getLogger(__name__)

# ── Styling ──
NAVY = "1B2A4A"
LIGHT_BLUE = "E8EEF4"
MED_BLUE = "C5D5E8"
LIGHT_GRAY = "F7F7F7"
MED_GRAY = "E0E0E0"
DARK_TEXT = "1A1A1A"
MUTED_TEXT = "666666"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Calibri", size=10, color=DARK_TEXT)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=NAVY)
CAMPAIGN_HDR_FILL = PatternFill("solid", fgColor=MED_BLUE)
CAMPAIGN_HDR_FONT = Font(name="Calibri", bold=True, size=10, color=NAVY)
TOTAL_FILL = PatternFill("solid", fgColor=MED_BLUE)
TOTAL_FONT = Font(name="Calibri", bold=True, size=10, color=DARK_TEXT)
ALT_ROW_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=11, color=MUTED_TEXT)
THIN_BORDER = Border(bottom=Side(style="thin", color=MED_GRAY),
                      right=Side(style="thin", color=MED_GRAY))
CR_HEADER_FILL = PatternFill("solid", fgColor=NAVY)
CR_TITLE_FONT = Font(name="Calibri", bold=True, size=22, color="FFFFFF")
CR_DATE_FONT = Font(name="Calibri", size=12, color="B0C4DE")

def write_to_excel(parsed_data_list, output_path, platform_name="",
                   inject_vba=True, excel_app=None):
    """Write the unified workbook: Search dashboard + Unified Data.

    Attempts to inject the interactive VBA search engine and save as .xlsm
    (Windows + Excel + trust setting required); otherwise leaves a plain
    .xlsx with the data intact. Returns (final_path, vba_error_or_None).
    excel_app: an open Excel COM Application to reuse for the injection
    (see engine.excel_vba.ExcelSession) — batch exports pay the Excel
    startup cost once instead of once per client.

    The workbook is ALWAYS rebuilt from scratch and the VBA re-injected.
    The old re-export path edited the existing .xlsm in place with
    openpyxl, which cannot round-trip macro workbooks: the ActiveX search
    box was silently dropped and the sheet/VBA wiring corrupted, so the
    search broke on every re-export of the same period. Existing rows are
    harvested from the previous file first, so re-export merge semantics
    (same-key replacement, distinct periods preserved) are unchanged.
    """
    base, _ext = os.path.splitext(output_path)
    xlsx_path = base + ".xlsx"
    xlsm_path = base + ".xlsm"
    source_path = None
    if os.path.exists(xlsm_path):
        source_path = xlsm_path        # the live report from a prior export
    elif os.path.exists(xlsx_path):
        source_path = xlsx_path

    # Harvest existing rows (read-only — the old workbook is never edited)
    existing_rows = []
    if source_path:
        try:
            old_wb = load_workbook(source_path, read_only=True, data_only=True)
            if "Unified Data" in old_wb.sheetnames:
                rows_iter = old_wb["Unified Data"].iter_rows(values_only=True)
                hdrs = next(rows_iter, None) or ()
                for values in rows_iter:
                    rd = {h: values[i] for i, h in enumerate(hdrs) if h}
                    if rd.get("metric_name"):
                        existing_rows.append(rd)
            old_wb.close()
        except Exception:
            logger.warning("Existing report at %s is unreadable — rebuilding "
                           "it from scratch", source_path, exc_info=True)
            existing_rows = []

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Collect and deduplicate
    new_rows = _collect_rows(parsed_data_list, platform_name)
    df = _build_dataframe(new_rows, existing_rows)

    if df.empty:
        raise ValueError(
            "No data to export — no rows matched the selected campaigns "
            "and date range.")

    # ── Unified Data sheet ──
    ws = wb.create_sheet("Unified Data")
    for ci, h in enumerate(SCHEMA_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center"); c.border = THIN_BORDER

    # itertuples over reindexed columns: iterrows built a Series per row and
    # made this loop the slowest part of large exports
    metric_ci = SCHEMA_COLUMNS.index("metric_value") + 1
    row_data = df.reindex(columns=SCHEMA_COLUMNS)
    for ri, values in enumerate(row_data.itertuples(index=False, name=None), 2):
        for ci, val in enumerate(values, 1):
            if pd.isna(val): val = None
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DATA_FONT; c.border = THIN_BORDER
            if ci == metric_ci and val is not None:
                _format_cell(c, val)

    widths = {"client": 18, "campaign": 35, "campaign_type": 15, "source": 18,
              "metric_level": 25, "metric_name": 20, "metric_value": 15,
              "start_date": 14, "end_date": 14}
    for ci, col in enumerate(SCHEMA_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 15)
    last_col = get_column_letter(len(SCHEMA_COLUMNS))
    # Pure-openpyxl polish: real Excel Table (filter buttons + banding),
    # frozen header, brand tab. A Table replaces auto_filter.
    from openpyxl.worksheet.table import Table, TableStyleInfo
    tbl = Table(displayName="UnifiedData", ref=f"A1:{last_col}{len(df) + 1}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                        showRowStripes=True)
    ws.add_table(tbl)
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "003057"

    # ── Search dashboard (sheet 0) + hidden index/config ──
    from engine.excel_search_dashboard import build_search_dashboard
    build_search_dashboard(wb, df)

    try:
        wb.save(xlsx_path)
    except PermissionError as e:
        raise PermissionError(
            f"Cannot write {xlsx_path} — the file is open in Excel. "
            "Close it and try again.") from e
    logger.info("Excel report written: %s (%d data rows)", xlsx_path, len(df))

    # ── Interactive search: inject VBA and upgrade to .xlsm ──
    final_path = xlsx_path
    vba_error = None
    if inject_vba:
        from engine.excel_vba import inject_search_vba
        final_path, vba_error = inject_search_vba(xlsx_path, app=excel_app)
        if vba_error:
            logger.warning("Search VBA not injected: %s", vba_error)

    # Never leave a stale macro workbook beside fresher data: if this
    # export ended as .xlsx (injection unavailable/failed) the old .xlsm
    # now holds OUTDATED rows — its data was harvested above, so drop it.
    if final_path.lower().endswith(".xlsx") and os.path.exists(xlsm_path):
        try:
            os.remove(xlsm_path)
        except OSError:
            logger.warning("Stale macro workbook is locked: %s", xlsm_path)
            stale_note = (
                f"A previous report ({os.path.basename(xlsm_path)}) is open "
                f"in Excel and shows OLD data — close it. The current data "
                f"is in {os.path.basename(final_path)}.")
            vba_error = f"{vba_error}\n\n{stale_note}" if vba_error else stale_note
    return final_path, vba_error
