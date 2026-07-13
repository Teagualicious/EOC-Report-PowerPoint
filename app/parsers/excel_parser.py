"""
Excel Parser with shared dictionary integration.
"""

import logging
from itertools import islice
import os
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from engine.errors import ParserError
from parsers.dictionary import (
    classify_columns, extract_row_context,
    build_level_rows, build_campaign_rows
)

logger = logging.getLogger(__name__)

_HEADER_SCAN_ROWS = 250


class ExcelParser:
    """Parses an Excel export (all sheets, smart header detection)."""

    def __init__(self, filepath):
        self.filepath = filepath

    def parse(self, build_outputs=True):
        """Parse the workbook. Raises ParserError if it cannot be read.

        ``build_outputs=False`` retains detected source rows but skips the
        automatic unified-row pass. This is faster when a platform mapping is
        about to rebuild those rows anyway.
        """
        if not os.path.exists(self.filepath):
            raise ParserError(f"File not found: {self.filepath}",
                              user_message="The file no longer exists at "
                                           f"{self.filepath}")
        try:
            wb = load_workbook(self.filepath, read_only=True, data_only=True)
        except (InvalidFileException, BadZipFile, KeyError) as e:
            raise ParserError(
                f"Invalid Excel file {self.filepath}: {e}",
                user_message="This is not a valid .xlsx or .xlsm file — was it "
                             "exported correctly? Legacy .xls files must first be "
                             "saved as .xlsx.") from e
        except OSError as e:
            raise ParserError(
                f"Cannot read {self.filepath}: {e}",
                user_message="The file could not be opened — is it open in "
                             "Excel? Close it and try again.") from e

        detected_tables = []
        level_data = []
        campaign_metrics = {}

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_iter = ws.iter_rows(values_only=True)
                # Header detection only needs the beginning of a report. Keeping
                # this bounded avoids materializing a million-row sheet twice.
                preview = [list(row) for row in islice(row_iter, _HEADER_SCAN_ROWS)]
                if len(preview) < 2:
                    continue

                best_header = self._find_header_row(preview)
                if best_header is None:
                    logger.debug("No table header found in first %d rows of %s",
                                 _HEADER_SCAN_ROWS, sheet_name)
                    continue

                raw_headers = preview[best_header]
                headers = [(j, str(c).strip()) for j, c in enumerate(raw_headers)
                           if c is not None and str(c).strip()]
                if len(headers) < 2:
                    continue

                header_names = [name for _, name in headers]
                header_cols = [idx for idx, _ in headers]
                classification = classify_columns(header_names)
                found_context = classification["context"]
                found_levels = classification["levels"]
                found_metrics = classification["metrics"]

                def data_rows():
                    yield from preview[best_header + 1:]
                    yield from row_iter

                table_rows = self._extract_data_rows_iter(
                    data_rows(), header_cols, header_names)
                if not table_rows:
                    continue

                detected_tables.append({
                    "headers": header_names,
                    "rows": table_rows,
                    "row_count": len(table_rows),
                    "sheet_name": sheet_name,
                    "found_levels": [(cn, ld["prefix"]) for cn, ld in found_levels],
                    "found_metrics": found_metrics,
                    "found_context": found_context,
                })

                if build_outputs:
                    if found_levels:
                        for data_row in table_rows:
                            ctx = extract_row_context(data_row, found_context)
                            level_data.extend(
                                build_level_rows(data_row, found_levels, found_metrics, ctx))
                    else:
                        for data_row in table_rows:
                            ctx = extract_row_context(data_row, found_context)
                            campaign_metrics.update(
                                build_campaign_rows(data_row, found_metrics, ctx))
        finally:
            wb.close()

        flat_metrics = {m["universal_name"]: m["value"]
                        for m in campaign_metrics.values()}
        campaign_name = next((m.get("campaign_name")
                              for m in campaign_metrics.values()
                              if m.get("campaign_name")), "")
        if not campaign_name:
            campaign_name = os.path.splitext(os.path.basename(self.filepath))[0]

        return {
            "source_file": os.path.basename(self.filepath),
            "campaign_name": campaign_name,
            "date_range": "See file",
            "metrics": flat_metrics,
            "detected_tables": detected_tables,
            "level_data": level_data,
            "campaign_metrics": campaign_metrics,
        }

    def _find_header_row(self, rows_data):
        best_header = None
        best_score = 0
        for i in range(len(rows_data) - 1):
            row = rows_data[i]
            if not row:
                continue
            str_cells = [c for c in row if c is not None and isinstance(c, str) and c.strip()]
            if len(str_cells) < 2:
                continue
            data_rows = 0
            for j in range(i + 1, min(i + 20, len(rows_data))):
                below = rows_data[j]
                if not below:
                    break
                numeric_count = 0
                non_empty = 0
                for ci, cell in enumerate(below):
                    if ci < len(row) and cell is not None:
                        non_empty += 1
                        if isinstance(cell, (int, float)):
                            numeric_count += 1
                        elif isinstance(cell, str) and self._looks_numeric(cell):
                            numeric_count += 1
                if non_empty >= 2 and numeric_count >= 1:
                    data_rows += 1
                elif non_empty == 0:
                    break
                else:
                    if data_rows > 0:
                        break
            score = len(str_cells) * data_rows
            if score > best_score:
                best_score = score
                best_header = i
        return best_header

    def _extract_data_rows(self, rows_data, best_header, header_cols, header_names):
        """Compatibility wrapper used by tests and older callers."""
        return self._extract_data_rows_iter(
            iter(rows_data[best_header + 1:]), header_cols, header_names)

    def _extract_data_rows_iter(self, rows, header_cols, header_names):
        table_rows = []
        consecutive_blanks = 0
        for raw_row in rows:
            row = list(raw_row)
            if not row or all(c is None for c in row):
                consecutive_blanks += 1
                if consecutive_blanks >= 2 and table_rows:
                    break
                continue

            row_vals = {}
            has_numeric = False
            has_any = False
            for col_idx, col_name in zip(header_cols, header_names):
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and str(val).strip():
                    has_any = True
                    cleaned = self._clean_value(val)
                    row_vals[col_name] = cleaned
                    if isinstance(cleaned, (int, float)):
                        has_numeric = True
                else:
                    row_vals[col_name] = ""

            if not has_any:
                consecutive_blanks += 1
                if consecutive_blanks >= 2 and table_rows:
                    break
                continue

            if not has_numeric and table_rows:
                text_vals = [str(v) for v in row if v is not None and str(v).strip()]
                long_text = any(len(text) > 50 for text in text_vals)
                row_text = " ".join(text_vals).lower()
                footer_keywords = any(keyword in row_text for keyword in
                    ("included", "downloaded", "report", "products",
                     "generated", "notes", "disclaimer"))
                if long_text or footer_keywords:
                    break

            consecutive_blanks = 0
            table_rows.append(row_vals)
        return table_rows

    def _looks_numeric(self, value_str):
        cleaned = str(value_str).replace("$", "").replace(",", "").replace("%", "").strip()
        try:
            float(cleaned)
            return True
        except (ValueError, TypeError):
            return False

    def _clean_value(self, value):
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            return value
        cleaned = str(value).strip()
        if not cleaned:
            return ""
        numeric = cleaned.replace("$", "").replace(",", "").replace("%", "").strip()
        try:
            if "." in numeric:
                return float(numeric)
            return int(numeric)
        except (ValueError, TypeError):
            return cleaned
